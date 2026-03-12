"""Excel 解析模块

使用 openpyxl 解析 .xlsx 文件。
每张 sheet 按 ROW_CHUNK_SIZE 行分组，生成 category="SheetRow" 元素。
行文本序列化格式：col1 | col2 | col3
"""
import logging
from pathlib import Path
from typing import List, TYPE_CHECKING

if TYPE_CHECKING:
    pass

from .models import ParsedDocument, ParsedElement

logger = logging.getLogger(__name__)

ROW_CHUNK_SIZE = 15  # 每组行数

try:
    import openpyxl  # noqa: F401
    OPENPYXL_AVAILABLE = True
except ImportError:
    openpyxl = None  # type: ignore[assignment]
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl 库不可用，无法解析 .xlsx 文件。请执行: pip install openpyxl")


class XlsxParser:
    """Excel 解析器

    支持 .xlsx 格式（不支持旧版 .xls，请另存为 .xlsx）。
    """

    def __init__(self, row_chunk_size: int = ROW_CHUNK_SIZE) -> None:
        self.row_chunk_size = row_chunk_size

    def parse(self, file_path: str) -> ParsedDocument:
        """解析 Excel 文件

        Args:
            file_path: .xlsx 文件路径

        Returns:
            ParsedDocument 对象

        Raises:
            RuntimeError: openpyxl 不可用
            ValueError: 传入 .xls 旧格式
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "openpyxl 库不可用，无法解析 Excel 文件。请执行: pip install openpyxl"
            )

        path = Path(file_path)
        if path.suffix.lower() == ".xls":
            raise ValueError(
                f"不支持旧版 .xls 格式: {path.name}。"
                "请在 Excel 中另存为 .xlsx 后重试。"
            )

        source_file = path.name
        elements = self._parse_with_openpyxl(file_path)

        return ParsedDocument(
            source_file=source_file,
            file_type="xlsx",
            elements=elements,
        )

    # ------------------------------------------------------------------
    # 内部方法
    # ------------------------------------------------------------------

    def _parse_with_openpyxl(self, file_path: str) -> List[ParsedElement]:
        """遍历所有 sheet，按行分组生成 ParsedElement 列表"""
        elements: List[ParsedElement] = []

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))

            if not all_rows:
                continue

            # 第一行作为列名（header）
            header_row = all_rows[0]
            column_names = [str(c) if c is not None else "" for c in header_row]
            col_str = ",".join(column_names)

            data_rows = all_rows[1:]
            if not data_rows:
                continue

            chunk_index = 0
            for start in range(0, len(data_rows), self.row_chunk_size):
                end = min(start + self.row_chunk_size, len(data_rows))
                chunk_rows = data_rows[start:end]

                # 序列化：header行 + 每行数据
                lines = [" | ".join(column_names)]
                for row in chunk_rows:
                    cells = [str(c) if c is not None else "" for c in row]
                    lines.append(" | ".join(cells))
                text = "\n".join(lines)

                if not text.strip():
                    continue

                elements.append(ParsedElement(
                    text=text,
                    category="SheetRow",
                    page_or_index=sheet_idx,
                    metadata={
                        "sheet_name": sheet_name,
                        "row_range": f"{start + 2}-{end + 1}",  # +1 因为有 header 行
                        "column_names": col_str,
                    },
                ))
                chunk_index += 1

        wb.close()
        return elements


def parse_xlsx(file_path: str) -> ParsedDocument:
    """解析 Excel 文件的便捷函数"""
    parser = XlsxParser()
    return parser.parse(file_path)

