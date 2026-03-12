"""DocumentRouter 测试套件

覆盖点：
- 各格式路由正确性（通过 mock 绕过真实文件）
- UnsupportedFileTypeError 抛出
- 懒加载不重复初始化
- IngestionPipeline 向后兼容 pdf_path 关键字参数
"""
import csv
import tempfile
from unittest.mock import patch, MagicMock

import pytest

from src.ingestion.document_router import DocumentRouter, UnsupportedFileTypeError, DocumentParseError
from src.ingestion.models import ParsedDocument, ParsedElement
from src.ingestion.plain_text_parser import PlainTextParser
from src.ingestion.xlsx_parser import XlsxParser


# ===========================================================================
# Fixtures
# ===========================================================================

@pytest.fixture
def router():
    return DocumentRouter()


def make_parsed_doc(file_type: str = "txt") -> ParsedDocument:
    """构造一个最小 ParsedDocument 用于 mock 返回值"""
    return ParsedDocument(
        source_file=f"test.{file_type}",
        file_type=file_type,  # type: ignore[arg-type]
        elements=[
            ParsedElement(text="hello", category="Text", page_or_index=0)
        ],
    )


# ===========================================================================
# DocumentRouter 路由测试（mock parse 方法，不需要真实文件）
# ===========================================================================

class TestDocumentRouterRouting:

    def test_route_pdf(self, router, tmp_path):
        """路由到 PDFParser，返回 file_type='pdf'"""
        fake_doc = make_parsed_doc("pdf")
        with patch("src.ingestion.pdf_parser.PDFParser.parse", return_value=fake_doc):
            doc = router.route(str(tmp_path / "report.pdf"))
        assert doc.file_type == "pdf"

    def test_route_docx(self, router, tmp_path):
        """路由到 DocxParser，返回 file_type='docx'"""
        fake_doc = make_parsed_doc("docx")
        with patch("src.ingestion.docx_parser.DocxParser.parse", return_value=fake_doc):
            doc = router.route(str(tmp_path / "report.docx"))
        assert doc.file_type == "docx"

    def test_route_pptx(self, router, tmp_path):
        """路由到 PptxParser，返回 file_type='pptx'"""
        fake_doc = make_parsed_doc("pptx")
        with patch("src.ingestion.pptx_parser.PptxParser.parse", return_value=fake_doc):
            doc = router.route(str(tmp_path / "slides.pptx"))
        assert doc.file_type == "pptx"

    def test_route_xlsx(self, router, tmp_path):
        """路由到 XlsxParser，返回 file_type='xlsx'"""
        fake_doc = make_parsed_doc("xlsx")
        with patch("src.ingestion.xlsx_parser.XlsxParser.parse", return_value=fake_doc):
            doc = router.route(str(tmp_path / "data.xlsx"))
        assert doc.file_type == "xlsx"

    @pytest.mark.parametrize("ext", [".txt", ".md", ".csv"])
    def test_route_plain_text(self, router, tmp_path, ext):
        """txt/md/csv 均路由到 PlainTextParser"""
        file_type = ext.lstrip(".")
        fake_doc = make_parsed_doc(file_type)
        with patch("src.ingestion.plain_text_parser.PlainTextParser.parse", return_value=fake_doc):
            doc = router.route(str(tmp_path / f"file{ext}"))
        assert doc.file_type == file_type

    def test_unsupported_extension(self, router, tmp_path):
        """不支持的扩展名抛出 UnsupportedFileTypeError"""
        with pytest.raises(UnsupportedFileTypeError) as exc_info:
            router.route(str(tmp_path / "file.xyz"))
        assert ".xyz" in str(exc_info.value)

    def test_unsupported_no_extension(self, router, tmp_path):
        """无扩展名文件抛出 UnsupportedFileTypeError"""
        with pytest.raises(UnsupportedFileTypeError):
            router.route(str(tmp_path / "noextfile"))

    def test_parse_error_wraps_exception(self, router, tmp_path):
        """底层解析异常被包装为 DocumentParseError"""
        with patch("src.ingestion.plain_text_parser.PlainTextParser.parse",
                   side_effect=Exception("broken file")):
            with pytest.raises(DocumentParseError) as exc_info:
                router.route(str(tmp_path / "bad.txt"))
        assert "broken file" in str(exc_info.value.__cause__)

    def test_lazy_load_no_double_init(self, tmp_path):
        """同一 Router 实例路由两次 PDF，PDFParser 只实例化一次"""
        router = DocumentRouter()
        fake_doc = make_parsed_doc("pdf")
        with patch("src.ingestion.pdf_parser.PDFParser.__init__", return_value=None) as mock_init, \
             patch("src.ingestion.pdf_parser.PDFParser.parse", return_value=fake_doc):
            router.route(str(tmp_path / "a.pdf"))
            router.route(str(tmp_path / "b.pdf"))
        # PDFParser.__init__ 只调用一次（懒加载缓存生效）
        assert mock_init.call_count == 1

    def test_plain_text_shared_instance(self, tmp_path):
        """txt / md / csv 共用同一个 PlainTextParser 实例"""
        router = DocumentRouter()
        fake_doc_txt = make_parsed_doc("txt")
        fake_doc_md = make_parsed_doc("md")
        with patch("src.ingestion.plain_text_parser.PlainTextParser.__init__", return_value=None) as mock_init, \
             patch("src.ingestion.plain_text_parser.PlainTextParser.parse",
                   side_effect=[fake_doc_txt, fake_doc_md]):
            router.route(str(tmp_path / "a.txt"))
            router.route(str(tmp_path / "b.md"))
        # PlainTextParser 只实例化一次
        assert mock_init.call_count == 1

    def test_supported_extensions_property(self, router):
        """supported_extensions 包含所有预期扩展名"""
        exts = router.supported_extensions
        for expected in [".pdf", ".docx", ".pptx", ".xlsx", ".txt", ".md", ".csv"]:
            assert expected in exts


# ===========================================================================
# PlainTextParser 实际解析测试（使用临时文件）
# ===========================================================================

class TestPlainTextParser:

    def test_parse_txt(self, tmp_path):
        f = tmp_path / "sample.txt"
        f.write_text("line1\nline2\nline3\n", encoding="utf-8")
        doc = PlainTextParser().parse(str(f))
        assert doc.file_type == "txt"
        assert len(doc.elements) >= 1
        assert doc.elements[0].category == "Text"
        assert "line_range" in doc.elements[0].metadata

    def test_parse_md(self, tmp_path):
        f = tmp_path / "readme.md"
        f.write_text("# Title\nsome content\n", encoding="utf-8")
        doc = PlainTextParser().parse(str(f))
        assert doc.file_type == "md"
        assert doc.elements[0].category == "Text"

    def test_parse_csv(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("name,value\nalice,1\nbob,2\n", encoding="utf-8")
        doc = PlainTextParser().parse(str(f))
        assert doc.file_type == "csv"
        assert doc.elements[0].category == "SheetRow"
        assert "column_names" in doc.elements[0].metadata
        assert "name" in doc.elements[0].metadata["column_names"]

    def test_txt_chunking(self, tmp_path):
        """超过 line_chunk_size 行时应产生多个 chunk"""
        lines = [f"line{i}\n" for i in range(70)]
        f = tmp_path / "big.txt"
        f.write_text("".join(lines), encoding="utf-8")
        doc = PlainTextParser(line_chunk_size=30).parse(str(f))
        assert len(doc.elements) == 3  # 70行 / 30 = 3 chunks (30+30+10)

    def test_csv_chunking(self, tmp_path):
        """超过 csv_row_chunk_size 行时应产生多个 chunk"""
        rows = [["col1", "col2"]] + [[f"v{i}a", f"v{i}b"] for i in range(20)]
        f = tmp_path / "big.csv"
        with open(str(f), "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerows(rows)
        doc = PlainTextParser(csv_row_chunk_size=15).parse(str(f))
        assert len(doc.elements) == 2  # 20行数据 / 15 = 2 chunks


# ===========================================================================
# XlsxParser 实际解析测试
# ===========================================================================

class TestXlsxParser:

    def test_parse_xlsx(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["项目", "本年", "上年"])
        for i in range(5):
            ws.append([f"item{i}", i * 100, i * 90])
        path = tmp_path / "test.xlsx"
        wb.save(str(path))

        doc = XlsxParser().parse(str(path))
        assert doc.file_type == "xlsx"
        assert len(doc.elements) >= 1
        assert doc.elements[0].category == "SheetRow"
        assert doc.elements[0].metadata["sheet_name"] == "Sheet1"
        assert "项目" in doc.elements[0].metadata["column_names"]

    def test_xls_raises_friendly_error(self, tmp_path):
        """传入 .xls 应抛出 ValueError 提示转换"""
        with pytest.raises(ValueError, match=".xls"):
            XlsxParser().parse(str(tmp_path / "old.xls"))

    def test_multi_sheet(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "资产"
        ws1.append(["科目", "金额"])
        ws1.append(["货币资金", 100])
        ws2 = wb.create_sheet("负债")
        ws2.append(["科目", "金额"])
        ws2.append(["短期借款", 50])
        path = tmp_path / "multi.xlsx"
        wb.save(str(path))

        doc = XlsxParser().parse(str(path))
        sheet_names = {e.metadata["sheet_name"] for e in doc.elements}
        assert "资产" in sheet_names
        assert "负债" in sheet_names


# ===========================================================================
# Pipeline 向后兼容测试
# ===========================================================================

class TestPipelineBackwardCompat:

    def _make_mock_pipeline(self):
        """创建一个跳过真实 embedding/storage 的 Pipeline mock"""
        from src.ingestion.pipeline import IngestionPipeline
        pipeline = IngestionPipeline.__new__(IngestionPipeline)
        from src.ingestion.document_router import DocumentRouter
        from src.ingestion.chunker import Chunker
        pipeline.router = DocumentRouter()
        pipeline.chunker = Chunker()

        # mock embedding 和 storage
        mock_embed = MagicMock()
        mock_embed.embed.return_value = [[0.1] * 10]
        mock_embed.get_dimension.return_value = 10
        pipeline.embedding_service = mock_embed

        mock_doc_store = MagicMock()
        mock_doc_store.count.return_value = 1
        pipeline.doc_store = mock_doc_store

        mock_vec_store = MagicMock()
        mock_vec_store.count.return_value = 1
        pipeline.vector_store = mock_vec_store

        return pipeline

    def test_pipeline_accepts_file_path(self, tmp_path):
        """pipeline.run(file_path=...) 正常工作"""
        f = tmp_path / "test.txt"
        f.write_text("hello world\n", encoding="utf-8")
        pipeline = self._make_mock_pipeline()
        result = pipeline.run(file_path=str(f))
        assert "file_path" in result
        assert result["chunk_count"] >= 1

    def test_pipeline_backward_compat_pdf_path(self, tmp_path):
        """旧版 pipeline.run(pdf_path=...) 关键字参数仍然有效"""
        f = tmp_path / "test.txt"
        f.write_text("hello world\n", encoding="utf-8")
        pipeline = self._make_mock_pipeline()
        result = pipeline.run(pdf_path=str(f))
        # 向后兼容：返回 dict 中保留 'pdf_path' key
        assert "pdf_path" in result

    def test_pipeline_run_no_args_raises(self):
        """不提供任何路径参数时应抛出 ValueError"""
        pipeline = self._make_mock_pipeline()
        with pytest.raises(ValueError):
            pipeline.run()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])


